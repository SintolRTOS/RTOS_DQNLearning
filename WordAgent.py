# -*- coding: utf-8 -*-
"""
Created on Mon Apr 29 18:26:17 2019

@author: wangjingyi
"""
import numpy as np
np.set_printoptions(suppress=True)
from openpyxl import load_workbook
import logging  # 引入logging模块
import time
import tensorflow as tf
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s')  # logging.basicConfig函数对日志的输出格式及方式做相关配置
logger.setLevel(level = logging.INFO)
handler = logging.FileHandler("WordGame/log/wordagent_log_" + str(time.time()) + '.txt')
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
console = logging.StreamHandler()
console.setLevel(logging.INFO)
logger.addHandler(console)

POPULARITY_BOUND = 1000000

class WordAgent(object):
    """docstring for ClassName"""
    def __init__(self,
                 filepath,
                 mode):
        super(WordAgent,self).__init__()
        self.filepath = filepath
        self.mode = mode
        self.parameter_size = 5
        self.result = []
        self.result_keywords = []
        self.select_index = 0;
        self.select_count = 0;
        self.select_process = 0.
        self.select_total = 0
        self.select_total_reward = 0.
    
    def openExcel(self):
        logging.info('start openpyxl openExcel!')
        self.wb = load_workbook(self.filepath)
        print('form sheetname: ' + self.wb.sheetnames[0])
        self.keytitle = self.wb.sheetnames[0];
        sheet = self.wb.get_sheet_by_name(str(self.keytitle))
        self.max_row = sheet.max_row
        self.max_column = sheet.max_column
        logging.info("form max_row: " + str(self.max_row))
        logging.info("form max_column: " + str(self.max_column))
        self.data = []

        for row in sheet.iter_rows(min_col=1, min_row=1, max_row=self.max_row, max_col=self.max_column):
            ROW = []
            for cell in row:
                ROW.append(cell.value)
            self.data.append(ROW)
        logging.info("all form data: ")
        logging.info(self.data)
        
    def reset(self):
        self.init_observation()
#        self.init_agent_obs()
        self.init_action_space()
        self.reward = 0.
        self.result.clear()
        self.result_keywords.clear()
        return self.observation
    
    def init_agent_obs(self):
        logging.info('init_game_obs-------------------------')
        self.select_index = 0;
        self.select_count = 0;
        self.select_process = 0.
        self.select_total = 3
        self.select_total_reward = 0.
        self.agent_obs = np.empty(8)
        self.agent_obs[0] = 0.
        self.agent_obs[1] = self.observation[0]
        self.agent_obs[2] = self.observation[1]
        self.agent_obs[3] = self.observation[2]
        self.agent_obs[4] = self.observation[3]
        self.agent_obs[5] = self.observation[4]
        self.agent_obs[6] = 0.
        self.agent_obs[7] = 0.
        
        
    def init_observation(self):
        logging.info('init_observation-------------------------')
        self.keywords_length = 10
#        self.keywords_length = len(self.data)
        logger.info('self.keywords_length: '+ str(self.keywords_length))
        self.observation = np.empty(self.keywords_length * self.parameter_size,float)
        popularity_max = self.data[0][4]
        popularity_min = self.data[0][4]
        conversion_max = self.data[0][5]
        conversion_min = self.data[0][5]
        transform_1_max = self.data[0][8]
        transform_1_min = self.data[0][8]
        transform_2_max = self.data[0][9]
        transform_2_min = self.data[0][9]
        transform_3_max = self.data[0][10]
        transform_3_min = self.data[0][10]
        for i in range(self.keywords_length):
            self.observation[i*self.parameter_size] = self.data[i][4]
            if self.observation[i*self.parameter_size] > popularity_max:
                popularity_max = self.observation[i*self.parameter_size]
            if self.observation[i*self.parameter_size] < popularity_min:
                popularity_min = self.observation[i*self.parameter_size]
                
            self.observation[i*self.parameter_size + 1] = self.data[i][5]
            if self.observation[i*self.parameter_size + 1] > conversion_max:
                conversion_max = self.observation[i*self.parameter_size + 1]
            if self.observation[i*self.parameter_size + 1] < conversion_min:
                conversion_min = self.observation[i*self.parameter_size + 1]
                
            self.observation[i*self.parameter_size + 2] = self.data[i][8]
            if self.observation[i*self.parameter_size + 2] > transform_1_max:
                transform_1_max = self.observation[i*self.parameter_size + 2]
            if self.observation[i*self.parameter_size + 2] < transform_1_min:
                transform_1_min = self.observation[i*self.parameter_size + 2]
                
            self.observation[i*self.parameter_size + 3] = self.data[i][9]
            if self.observation[i*self.parameter_size + 3] > transform_2_max:
                transform_2_max = self.observation[i*self.parameter_size + 3]
            if self.observation[i*self.parameter_size + 3] < transform_2_min:
                transform_2_min = self.observation[i*self.parameter_size + 3]
                
            self.observation[i*self.parameter_size + 4] = self.data[i][10]
            if self.observation[i*self.parameter_size + 4] > transform_3_max:
                transform_3_max = self.observation[i*self.parameter_size + 4]
            if self.observation[i*self.parameter_size + 4] < transform_3_min:
                transform_3_min = self.observation[i*self.parameter_size + 4]
#            self.observation[i*self.parameter_size + 5] = i
#            logger.info('----------------self.observation i: ' + str(i))
#            logger.info('self.observation[i*self.parameter_size]: ' + str(self.observation[i*self.parameter_size]))
#            logger.info('self.observation[i*self.parameter_size + 1]: ' + str(self.observation[i*self.parameter_size + 1]))
#            logger.info('self.observation[i*self.parameter_size + 2]: ' + str(self.observation[i*self.parameter_size + 2]))
#            logger.info('self.observation[i*self.parameter_size + 3]: ' + str(self.observation[i*self.parameter_size + 3]))
#            logger.info('self.observation[i*self.parameter_size + 4]: ' + str(self.observation[i*self.parameter_size + 4]))
#            logger.info('----------------end---------------- ')
#            ROW = []
#            ROW.append(self.data[i][5])
#            ROW.append(self.data[i][6])
#            ROW.append(self.data[i][9])
#            ROW.append(self.data[i][10])
#            ROW.append(self.data[i][11])
#            self.observation.append(ROW)
        #normal action
        normal_check_popularity = 0.
        normal_check_conversion = 0.
        normal_check_transform_1 = 0.
        normal_check_transform_2 = 0.
        normal_check_transform_3 = 0.
        logger.debug('popularity_max:' + str(popularity_max))
        logger.debug('popularity_min:' + str(popularity_min))
        logger.debug('conversion_max:' + str(conversion_max))
        logger.debug('conversion_min:' + str(conversion_min))
        logger.debug('transform_1_max:' + str(transform_1_max))
        logger.debug('transform_1_min:' + str(transform_1_min))
        logger.debug('transform_2_max:' + str(transform_2_max))
        logger.debug('transform_2_min:' + str(transform_2_min))
        logger.debug('transform_3_max:' + str(transform_3_max))
        logger.debug('transform_3_min:' + str(transform_3_min))
        for j in range(self.keywords_length):
            self.observation[j*self.parameter_size] = (self.observation[j*self.parameter_size] - popularity_min) / (popularity_max - popularity_min)
            normal_check_popularity+=self.observation[j*self.parameter_size]
            
            self.observation[j*self.parameter_size + 1] = (self.observation[j*self.parameter_size + 1] - conversion_min) / (conversion_max - conversion_min)
            normal_check_conversion+=self.observation[j*self.parameter_size + 1]
            
            self.observation[j*self.parameter_size + 2] = (self.observation[j*self.parameter_size + 2] - transform_1_min) / (transform_1_max - transform_1_min)
            normal_check_transform_1+=self.observation[j*self.parameter_size + 2]
            
            self.observation[j*self.parameter_size + 3] = (self.observation[j*self.parameter_size + 3] - transform_2_min) / (transform_2_max - transform_2_min)
            normal_check_transform_2+=self.observation[j*self.parameter_size + 3]
            
            self.observation[j*self.parameter_size + 4] = (self.observation[j*self.parameter_size + 4] - transform_3_min) / (transform_3_max - transform_3_min)
            normal_check_transform_3+=self.observation[j*self.parameter_size + 4]
            
            logger.debug('----------------self.observation i: ' + str(j))
            logger.debug('self.observation[i*self.parameter_size]: ' + str(self.observation[j*self.parameter_size]))
            logger.debug('self.observation[i*self.parameter_size + 1]: ' + str(self.observation[j*self.parameter_size + 1]))
            logger.debug('self.observation[i*self.parameter_size + 2]: ' + str(self.observation[j*self.parameter_size + 2]))
            logger.debug('self.observation[i*self.parameter_size + 3]: ' + str(self.observation[j*self.parameter_size + 3]))
            logger.debug('self.observation[i*self.parameter_size + 4]: ' + str(self.observation[j*self.parameter_size + 4]))
            logger.debug('----------------end---------------- ')
            
        logger.debug(str(self.observation))
        logger.debug('normal_check_popularity:' + str(normal_check_popularity))
        logger.debug('normal_check_conversion:' + str(normal_check_conversion))
        logger.debug('normal_check_transform_1:' + str(normal_check_transform_1))
        logger.debug('normal_check_transform_2:' + str(normal_check_transform_2))
        logger.debug('normal_check_transform_3:' + str(normal_check_transform_3))
        logger.info('end_observation-------------------------')
        return self.observation
    
    def get_observation(self):
        return self.observation
    
#    def get_agent_obs(self):
#        return self.agent_obs
    
    def init_action_space(self):
        self.action_space = [2.,]
    
    def get_action_space(self):
        return self.action_space
    
    def get_result_keywords(self):
        logger.info('--------------get_result_keywords----------------')
        self.result_keywords.clear()
        count = len(self.result)
        self.result.sort()
        logger.info('self.result: ' + str(self.result))
        logger.info('result count: ' + str(count))
        for i in range(count):
            key_id = int(self.result[i])
            key_words = self.data[key_id][1]
            self.result_keywords.append(key_words)
            logger.info('key_id: ' + str(key_id))
            logger.info('keywords: ' + str(key_words))
        
        logger.info('result keywords: ' + str(self.result_keywords))
        logger.info('--------------end get_result_keywords----------------')
        return self.result_keywords
    
    def step_agent(self,u):
        logger.info('step_agent select: ' + str(u))
        a_value =  float(u[0])
        isselect = False
        if a_value > 0:
            isselect = True
        logger.info('step_agent isselect: ' + str(isselect))
        index = self.select_index
        logger.info('step_agent----------------------:' + str(index))
        popularity = self.observation[index*self.parameter_size]
        conversion = self.observation[index*self.parameter_size + 1]
        transform_1 = self.observation[index*self.parameter_size + 2]
        transform_2 = self.observation[index*self.parameter_size + 3]
        transform_3 = self.observation[index*self.parameter_size + 4]
        keywords_id = index
        
        self.select_index += 1
        self.reward = 0
        if isselect:
            #add mistake error value
            self.observation[index*self.parameter_size] = -abs(popularity)
            self.reward = popularity*conversion + popularity*transform_1*2 + popularity*transform_2 + popularity*transform_3
            self.result.append(int(keywords_id))
            self.select_count += 1
            self.select_total_reward += self.reward
        
        
        logger.info('step_agent self.reward: ' + str(self.reward))
        logger.info('step_agent self.result: ' + str(self.result))
        logger.info('step_agent self.select_total_reward: ' + str(self.select_total_reward))
        self.select_process = float(self.select_count) / float(self.select_total)
        index += 1
        popularity = self.observation[index*self.parameter_size]
        conversion = self.observation[index*self.parameter_size + 1]
        transform_1 = self.observation[index*self.parameter_size + 2]
        transform_2 = self.observation[index*self.parameter_size + 3]
        transform_3 = self.observation[index*self.parameter_size + 4]
        self.agent_obs[0] = float(self.select_index) / float(self.keywords_length)
        self.agent_obs[1] = popularity
        self.agent_obs[2] = conversion
        self.agent_obs[3] = transform_1
        self.agent_obs[4] = transform_2
        self.agent_obs[5] = transform_3
        self.agent_obs[6] = self.select_process
        
        idDone = ((self.select_count == self.select_total) or (index == (self.keywords_length - 1)))
        logger.info('step_agent idDone: ' + str(idDone))
        logger.info('end step_agent----------------------')
        return self.get_agent_obs(),self.reward,idDone,{}
        
    
    def step(self,u):
        logger.info('WordAgent select: ' + str(u))
        a_value =  float(u[0])
        pervalue = (a_value + 1.) / 2
        logger.info('WordAgent pervalue: ' + str(pervalue))
        index = int(pervalue * float(self.keywords_length - 1))
        logger.info('--------------step----------------')
        logger.info('WordAgent select index: ' + str(index))
        
        logger.info('step self.observation index: ' + str(index))
        popularity = self.observation[index*self.parameter_size]
        logger.info('step popularity: ' + str(popularity))
        conversion = self.observation[index*self.parameter_size + 1]
        transform_1 = self.observation[index*self.parameter_size + 2]
        transform_2 = self.observation[index*self.parameter_size + 3]
        transform_3 = self.observation[index*self.parameter_size + 4]
        keywords_id = index
        #add mistake error value
        self.observation[index*self.parameter_size] = -abs(popularity)
        self.reward = popularity*conversion + popularity*transform_1*2 + popularity*transform_2 + popularity*transform_3
        logger.info('WordAgent self.reward: ' + str(self.reward))
        self.result.append(int(keywords_id))

        logger.info('step self.observation[index*self.parameter_size]: ' + str(self.observation[index*self.parameter_size]))
        logger.info('step self.observation[index*self.parameter_size + 1]: ' + str(self.observation[index*self.parameter_size + 1]))
        logger.info('step self.observation[index*self.parameter_size + 2]: ' + str(self.observation[index*self.parameter_size + 2]))
        logger.info('step self.observation[index*self.parameter_size + 3]: ' + str(self.observation[index*self.parameter_size + 3]))
        logger.info('step self.observation[index*self.parameter_size + 4]: ' + str(self.observation[index*self.parameter_size + 4]))
#        logger.info('step self.observation[index*self.parameter_size + 5]: ' + str(self.observation[index*self.parameter_size + 5]))
        logger.info('step self.observation result: ' + str(self.result))
        logger.info('----------------step end---------------- ')
        
        return self.get_observation(),self.reward,False,{}
        
            
        
        
#
#print('test openpyxl sucessful!')
#agent = WordAgent('assert/keyword.xlsx','xlsx')
#agent.openExcel()
#agent.reset()
        
#info = np.array([1,2,3,4,5,6,7,8,9,0])
#print(info.shape)
#observation =  np.delete(info,
#                  (3,4,5),
#                   axis = 0)
#print(observation.shape)

    
        